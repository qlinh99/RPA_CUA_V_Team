# -*- coding: utf-8 -*-
"""
Backend ĐÍCH = EXCEL: coi 1 sheet báo cáo như "form", dòng header = các trường,
mỗi chứng từ = 1 dòng thêm mới. Ghi thẳng bằng openpyxl (Bậc 1 — không mở Excel).

  inspect_excel(path, sheet, header_row) -> (sheet_name, fields)
  append_row(path, sheet, header_row, fields, values_by_id) -> số dòng vừa ghi
  make_template(path) -> tạo file báo cáo mẫu để test
"""
from __future__ import annotations
import re
from datetime import datetime
from pathlib import Path

HEADERS_MAU = [
    "Số hoá đơn", "Ký hiệu hoá đơn", "Ngày lập", "Tên nhà cung cấp",
    "MST nhà cung cấp", "Diễn giải", "Tiền trước thuế", "Thuế suất GTGT",
    "Tổng tiền thanh toán",
]

# Trường hóa đơn BẮT BUỘC phải có giá trị (theo nhãn cột, chuẩn hoá thường)
_REQUIRED_INVOICE_LABELS: "set[str]" = {
    "số hoá đơn", "ký hiệu hoá đơn", "ngày lập", "ngày hoá đơn",
    "tên nhà cung cấp", "mst nhà cung cấp", "mã số thuế nhà cung cấp",
    "tiền trước thuế", "tổng tiền thanh toán", "tổng thanh toán",
}

# Hint phân biệt trường dễ nhầm (nhãn thường → hint)
_FIELD_HINTS: "dict[str, str]" = {
    "số hoá đơn":      "Con số tuần tự trên hoá đơn, thường chỉ gồm chữ số như '0000123' — KHÁC với Ký hiệu",
    "ký hiệu hoá đơn": "Mã định danh mẫu/ký hiệu, thường dạng chữ+số như '1C25T', 'K25TSN' — KHÁC với Số hoá đơn",
}


def _guess_type(label: str) -> str:
    l = (label or "").lower()
    if any(k in l for k in ("ngày", "ngay", "date")):
        return "date"
    # [E3] Nhận dạng cột số tiền → type "number" để _coerce chuyển đúng
    if any(k in l for k in ("tiền", "tien", "thuế", "thue", "giá", "gia",
                             "phí", "phi", "tổng", "tong", "amount",
                             "price", "tax", "total")):
        return "number"
    return "text"


def inspect_excel(path: str, sheet: "str | None" = None, header_row: int = 1):
    """Đọc dòng header -> danh sách trường [{id, label, type, col}]."""
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    fields = []
    for col, cell in enumerate(ws[header_row], 1):
        label = "" if cell.value is None else str(cell.value).strip()
        if not label:
            continue
        label_lower = label.lower().strip()
        fields.append({
            "id": f"col{col}",
            "label": label,
            "type": _guess_type(label),
            "col": col,
            "required": label_lower in _REQUIRED_INVOICE_LABELS,
            "hint": _FIELD_HINTS.get(label_lower, ""),
        })
    name = ws.title
    wb.close()
    return name, fields


def _coerce(value, ftype: str):
    """Chuyển chuỗi OCR thành kiểu phù hợp để Excel hiểu (số/ngày)."""
    if value in (None, ""):
        return None
    if ftype == "date":
        for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(str(value), fmt).date()
            except ValueError:
                pass
        return value
    # [E3] Cột số tiền: bỏ ký tự phân cách, giữ nguyên phần nguyên (VND không có thập phân)
    if ftype == "number":
        digits = re.sub(r"[^\d]", "", str(value))
        return int(digits) if digits else None
    s = str(value)
    if re.fullmatch(r"\d{1,15}", s):
        # GIỮ text nếu có số 0 đứng đầu (số hoá đơn/MST) — tránh mất '0'
        if len(s) > 1 and s.startswith("0"):
            return s
        return int(s)               # số tiền -> int để Excel cộng được
    return value


def _true_last_row(ws, header_row: int) -> int:
    """Tìm dòng cuối thực sự có dữ liệu — scan ngược từ max_row qua TẤT CẢ cột."""
    max_col = ws.max_column or 1
    for r in range(ws.max_row, header_row, -1):
        if any(ws.cell(row=r, column=c).value is not None
               for c in range(1, max_col + 1)):
            return r
    return header_row


def _find_inv_col(fields: "list[dict]") -> "int | None":
    """Trả chỉ số cột có nhãn chứa 'số hoá đơn' / 'invoice' (dùng kiểm tra trùng)."""
    for f in fields:
        if any(k in f["label"].lower()
               for k in ("số hoá đơn", "so hoa don", "invoice no", "invoice number")):
            return f["col"]
    return None


def append_row(path: str, sheet: "str | None", header_row: int,
               fields: "list[dict]", values_by_id: dict) -> int:
    """Thêm 1 dòng mới ngay sau dữ liệu hiện có, ghi từng giá trị vào đúng cột."""
    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb[sheet] if sheet else wb.active

    # [E2] Tìm dòng cuối thực sự (scan tất cả cột, tránh max_row sai khi có ô trống)
    row = _true_last_row(ws, header_row) + 1

    # [E4] Kiểm tra trùng số hoá đơn trước khi ghi
    inv_col = _find_inv_col(fields)
    if inv_col:
        inv_id = next((f["id"] for f in fields if f["col"] == inv_col), None)
        inv_val = _coerce(values_by_id.get(inv_id, ""), "text") if inv_id else None
        if inv_val:
            for r_idx in range(header_row + 1, row):
                if str(ws.cell(row=r_idx, column=inv_col).value or "") == str(inv_val):
                    print(f"   ⚠️  TRÙNG: Số hoá đơn {inv_val!r} đã có ở dòng {r_idx}"
                          f" — vẫn thêm dòng mới (kiểm tra lại nếu không muốn trùng)")
                    break

    for f in fields:
        val = _coerce(values_by_id.get(f["id"]), f["type"])
        if val is not None:
            ws.cell(row=row, column=f["col"], value=val)
    wb.save(path)

    # [E1] Read-back: xác nhận ô thực sự có dữ liệu sau khi save
    wb_v = load_workbook(path, read_only=True, data_only=True)
    ws_v = wb_v[ws.title]
    null_fields = [f["label"] for f in fields
                   if values_by_id.get(f["id"]) not in (None, "")
                   and ws_v.cell(row=row, column=f["col"]).value is None]
    wb_v.close()
    if null_fields:
        print(f"   ⚠️  Read-back: {len(null_fields)} ô ghi nhưng đọc lại null: "
              + ", ".join(null_fields[:4]))

    return row


def make_template(path: str) -> None:
    """Tạo file báo cáo tài chính mẫu (chỉ có dòng header)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = "BaoCao"
    for col, h in enumerate(HEADERS_MAU, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="5645D4")
        ws.column_dimensions[c.column_letter].width = max(14, len(h) + 2)
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"✅ Đã tạo báo cáo mẫu: {path}")


if __name__ == "__main__":
    import sys
    out = sys.argv[1] if len(sys.argv) > 1 else "bao_cao_tai_chinh.xlsx"
    make_template(out)
