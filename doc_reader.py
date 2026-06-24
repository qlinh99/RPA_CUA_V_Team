# -*- coding: utf-8 -*-
"""
Đọc chứng từ thành danh sách "trang" để đưa vào OCR / LLM trích xuất.
Dùng chung cho mọi tool. Kế thừa chiến thuật lai của covert.py:
  - PDF điện tử (có text) -> lấy text bằng pdfplumber (rẻ, chính xác)
  - PDF scan / ảnh        -> render/encode ảnh cho LLM vision OCR
  - File DỮ LIỆU (Excel/CSV/Word/text...) -> đọc thẳng thành markdown/text

Trả về list các tuple:
  ('text',  markdown)        # PDF điện tử, Excel, CSV, Word, txt...
  ('image', b64, (w,h))      # ảnh / PDF scan
"""
from __future__ import annotations
from pathlib import Path

from test_image_processing import process_image, prepare_for_api

IMG_EXT = {".jpg", ".jpeg", ".png", ".bmp", ".webp", ".tif", ".tiff"}
CSV_EXT = {".csv", ".tsv"}
XLSX_EXT = {".xlsx", ".xlsm"}          # openpyxl đọc trực tiếp
DOCX_EXT = {".docx"}                    # đọc bằng zipfile (không cần lib ngoài)
TEXT_EXT = {".txt", ".md", ".markdown", ".json", ".log"}
OFFICE_LEGACY = {".xls", ".doc"}       # định dạng cũ -> cần COM Office
PDF_TEXT_MIN = 50  # >50 ký tự text => PDF điện tử
MAX_ROWS = 300     # giới hạn dòng bảng (Excel/CSV) để tránh prompt quá dài


def _table_to_markdown(table_data) -> str:
    if not table_data or not any(table_data):
        return ""
    rows = [r for r in table_data if r and any(c is not None for c in r)]
    if not rows:
        return ""
    md = ""
    for i, row in enumerate(rows):
        clean = [str(c).replace("\n", "<br>").strip() if c is not None else "" for c in row]
        md += "| " + " | ".join(clean) + " |\n"
        if i == 0:
            md += "| " + " | ".join(["---"] * len(clean)) + " |\n"
    return md + "\n"


def _pdf_page_text(page_plumber) -> str:
    table_bboxes = [t.bbox for t in page_plumber.find_tables()]

    def not_within_table(obj):
        if obj["object_type"] != "char":
            return True
        for x0, top, x1, bottom in table_bboxes:
            if (x0 <= obj["x0"] <= x1) and (top <= obj["top"] <= bottom):
                return False
        return True

    clean_text = page_plumber.filter(not_within_table).extract_text() or ""
    md_tables = "".join(_table_to_markdown(t) for t in page_plumber.extract_tables())
    return (clean_text + "\n\n" + md_tables).strip()


# ── file DỮ LIỆU -> markdown/text ────────────────────────────────────────────
def _rows_to_markdown(rows: "list[list]") -> str:
    """Danh sách dòng -> bảng markdown (dòng đầu = header). Bỏ dòng rỗng, cắt MAX_ROWS."""
    clean = [[("" if c is None else str(c)).replace("\n", "<br>").strip() for c in r]
             for r in rows if r and any(c not in (None, "") for c in r)]
    if not clean:
        return ""
    if len(clean) > MAX_ROWS:
        clean = clean[:MAX_ROWS] + [[f"... (còn {len(rows) - MAX_ROWS} dòng nữa, đã cắt)"]]
    width = max(len(r) for r in clean)
    md = ""
    for i, r in enumerate(clean):
        r = r + [""] * (width - len(r))
        md += "| " + " | ".join(r) + " |\n"
        if i == 0:
            md += "| " + " | ".join(["---"] * width) + " |\n"
    return md


def _csv_to_text(path: str) -> str:
    import csv
    delim = "\t" if Path(path).suffix.lower() == ".tsv" else ","
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        sample = f.read(2048)
        f.seek(0)
        if delim == ",":
            try:
                delim = csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
            except Exception:
                delim = ","
        rows = list(csv.reader(f, delimiter=delim))
    return _rows_to_markdown(rows)


def _xlsx_to_text(path: str) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    parts = []
    for ws in wb.worksheets:
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        md = _rows_to_markdown(rows)
        if md.strip():
            parts.append(f"### Sheet: {ws.title}\n{md}")
    wb.close()
    return "\n\n".join(parts)


def _docx_to_text(path: str) -> str:
    """Đọc .docx KHÔNG cần lib ngoài: bóc text trong word/document.xml."""
    import re
    import zipfile
    with zipfile.ZipFile(path) as z:
        xml = z.read("word/document.xml").decode("utf-8", errors="replace")
    xml = re.sub(r"</w:p>", "\n", xml)            # mỗi đoạn = 1 dòng
    xml = re.sub(r"<w:tab[^>]*/>", "\t", xml)
    text = re.sub(r"<[^>]+>", "", xml)            # bỏ mọi thẻ XML
    import html
    return html.unescape("\n".join(ln.rstrip() for ln in text.splitlines() if ln.strip()))


def _office_legacy_to_text(path: str) -> str:
    """.xls/.doc (định dạng cũ) -> dùng COM Office (cần cài MS Office)."""
    import win32com.client as win32  # type: ignore
    ext = Path(path).suffix.lower()
    full = str(Path(path).resolve())
    if ext == ".xls":
        app = win32.Dispatch("Excel.Application"); app.Visible = False
        try:
            wb = app.Workbooks.Open(full, ReadOnly=True)
            parts = []
            for ws in wb.Worksheets:
                ur = ws.UsedRange
                rows = ur.Value or []
                if isinstance(rows, tuple):
                    rows = [list(r) if isinstance(r, tuple) else [r] for r in rows]
                md = _rows_to_markdown(rows)
                if md.strip():
                    parts.append(f"### Sheet: {ws.Name}\n{md}")
            wb.Close(SaveChanges=False)
            return "\n\n".join(parts)
        finally:
            app.Quit()
    else:  # .doc
        app = win32.Dispatch("Word.Application"); app.Visible = False
        try:
            d = app.Documents.Open(full, ReadOnly=True)
            txt = d.Content.Text
            d.Close(SaveChanges=False)
            return txt
        finally:
            app.Quit()


def file_to_pages(path: str) -> "list[tuple]":
    ext = Path(path).suffix.lower()

    if ext in IMG_EXT:
        b64, _q, size = process_image(path)
        return [("image", b64, size)]

    # ── file dữ liệu: đọc thẳng thành text/markdown ──
    if ext in TEXT_EXT:
        print(f"  📄 File text ({ext}) → đọc thẳng")
        return [("text", Path(path).read_text(encoding="utf-8", errors="replace"))]
    if ext in CSV_EXT:
        print(f"  📊 File {ext} → bảng markdown")
        return [("text", _csv_to_text(path))]
    if ext in XLSX_EXT:
        print(f"  📊 Excel ({ext}) → bảng markdown (mọi sheet)")
        return [("text", _xlsx_to_text(path))]
    if ext in DOCX_EXT:
        print("  📝 Word (.docx) → text")
        return [("text", _docx_to_text(path))]
    if ext in OFFICE_LEGACY:
        print(f"  📝 Office cũ ({ext}) → COM Office (cần cài MS Office)")
        return [("text", _office_legacy_to_text(path))]

    if ext == ".pdf":
        import fitz
        import pdfplumber
        import cv2
        import numpy as np

        pages: list[tuple] = []
        doc = fitz.open(path)
        plumber = pdfplumber.open(path)
        for i in range(len(doc)):
            txt = _pdf_page_text(plumber.pages[i])
            if len(txt.strip()) > PDF_TEXT_MIN:
                print(f"  ⚡ Trang {i+1}/{len(doc)}: PDF điện tử → text thẳng (không gửi ảnh)")
                pages.append(("text", txt))
            else:
                print(f"  📸 Trang {i+1}/{len(doc)}: PDF scan → render ảnh + LLM OCR")
                pix = doc[i].get_pixmap(matrix=fitz.Matrix(2, 2))
                arr = np.frombuffer(pix.tobytes("jpg"), np.uint8)
                img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
                b64, size = prepare_for_api(img)
                pages.append(("image", b64, size))
        plumber.close()
        doc.close()
        return pages

    raise ValueError(
        f"Định dạng không hỗ trợ: {ext}\n"
        "  Nhận: ảnh (.jpg/.png/...), .pdf, .xlsx/.xlsm/.xls, .csv/.tsv, "
        ".docx/.doc, .txt/.md/.json")
